"""
Builds the single colour-coded Excel workbook from processed dataframes.
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import config

HEADER_FILL = PatternFill(start_color=config.NAVY, end_color=config.NAVY, fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def style_of_remark(remark):
    if remark == "TRUE":
        return config.GREEN, config.GREEN_FONT
    if remark == "Stock Mismatch" or remark == "Mismatch":
        return config.RED, config.RED_FONT
    if remark in ("UPDATE 0", "UPDATE STOCK"):
        return config.ORANGE, config.ORANGE_FONT
    if remark in config.ALL_NOT_FOUND_LABELS:
        return config.GREY, config.GREY_FONT
    return None, None


def write_df_sheet(wb, sheet_name, df, remark_col="Remark", status_col="Status_Result"):
    ws = wb.create_sheet(sheet_name[:31])
    ws.sheet_view.showGridLines = False

    cols = list(df.columns)
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    remark_idx = cols.index(remark_col) + 1 if remark_col in cols else None
    status_idx = cols.index(status_col) + 1 if status_col and status_col in cols else None

    for i, (_, r) in enumerate(df.iterrows(), start=2):
        remark_val = r.get(remark_col, None) if remark_col in cols else None
        fill_color, font_color = style_of_remark(remark_val) if remark_val is not None else (None, None)
        row_fill = PatternFill(start_color=config.ALT_ROW, end_color=config.ALT_ROW, fill_type="solid") if i % 2 == 0 else None

        for j, col in enumerate(cols, start=1):
            val = r[col]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if row_fill is not None:
                cell.fill = row_fill

        if remark_idx and fill_color:
            idxs = [remark_idx] + ([status_idx] if status_idx else [])
            for idx in idxs:
                c = ws.cell(row=i, column=idx)
                c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                c.font = Font(name="Arial", size=10, bold=True, color=font_color)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for j, col in enumerate(cols, start=1):
        sample = df[col].astype(str).values[:200]
        max_len = max([len(str(col))] + [len(str(v)) for v in sample]) if len(sample) else len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 10), 40)

    return ws


def write_summary_sheet(wb, marketplaces, kpi_results, extra_notes=None):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Multi-Marketplace Stock Validation"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=config.NAVY)
    ws["A2"] = datetime.now().strftime("Generated: %d %b %Y, %H:%M")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")

    row = 4
    metric_order = ["Total SKUs", "Matched", "Stock Mismatch", "UPDATE 0", "UPDATE STOCK",
                     "NOT FOUND", "Total Issues", "Accuracy %", "Not In Product Master"]

    for name, df, label in marketplaces:
        ws.cell(row=row, column=1, value=f"{name.upper()} — StockValidation vs Marketplace Stock")
        ws.cell(row=row, column=1).font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

        kpis = kpi_results[name]
        for metric in metric_order:
            if metric not in kpis:
                continue
            val = kpis[metric]
            label_metric = label if metric == "NOT FOUND" else metric
            ws.cell(row=row, column=1, value=label_metric).font = BODY_FONT
            cell_val = ws.cell(row=row, column=2, value=val)
            cell_val.font = Font(name="Arial", size=10, bold=True)
            cell_val.alignment = Alignment(horizontal="right")

            fill_color = None
            if metric == "Matched":
                fill_color = config.GREEN
            elif metric in ("Stock Mismatch", "Total Issues"):
                fill_color = config.RED
            elif metric in ("UPDATE 0", "UPDATE STOCK"):
                fill_color = config.ORANGE
            elif metric in ("NOT FOUND", "Not In Product Master"):
                fill_color = config.GREY
            if fill_color:
                cell_val.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            if metric == "Accuracy %":
                cell_val.number_format = '0.00"%"'
            row += 1
        row += 1

    if extra_notes:
        ws.cell(row=row, column=1, value="Notes").font = Font(name="Arial", size=11, bold=True, color=config.NAVY)
        row += 1
        for note in extra_notes:
            ws.cell(row=row, column=1, value=f"• {note}").font = Font(name="Arial", size=9, italic=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14


def build_workbook(
    marketplaces,
    kpi_results,
    soh_vs_all_df=None,
    master_coverage_df=None,
    notes=None,
) -> io.BytesIO:
    """
    marketplaces: list of (name, df, not_found_label) tuples.
    kpi_results: {name: kpi_dict}
    soh_vs_all_df: optional DataFrame from validate.soh_vs_all()
    master_coverage_df: optional DataFrame from validate.master_coverage_report()
    notes: optional list of strings shown on the Summary sheet
    """
    wb = Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, marketplaces, kpi_results, notes)

    for name, df, label in marketplaces:
        write_df_sheet(wb, f"{name}_StockVal", df)
        mism = df[df["Remark"] != "TRUE"]
        if len(mism) > 0:
            write_df_sheet(wb, f"{name}_Mismatches", mism)

    sheet_order = ["Summary"]
    for name, df, label in marketplaces:
        sheet_order.append(f"{name}_StockVal"[:31])
        mname = f"{name}_Mismatches"[:31]
        if mname in wb.sheetnames:
            sheet_order.append(mname)

    if soh_vs_all_df is not None and not soh_vs_all_df.empty:
        write_df_sheet(wb, "SOH_vs_ALL", soh_vs_all_df, remark_col="Remark", status_col=None)
        sheet_order.append("SOH_vs_ALL")
        mism = soh_vs_all_df[soh_vs_all_df["Remark"] != "TRUE"]
        if len(mism) > 0:
            write_df_sheet(wb, "SOH_Mismatches", mism, remark_col="Remark", status_col=None)
            sheet_order.append("SOH_Mismatches")

    if master_coverage_df is not None and not master_coverage_df.empty:
        write_df_sheet(wb, "Product_Master_Coverage", master_coverage_df, remark_col=None, status_col=None)
        sheet_order.append("Product_Master_Coverage")

    wb._sheets = [wb[s] for s in sheet_order]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
