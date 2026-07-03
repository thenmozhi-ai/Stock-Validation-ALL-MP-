"""
Stock Validation ALL MP — Streamlit app
Validates Expected Stock (from StockValidation reports) against live marketplace
stock files for Lazada, Shopee, TikTok, and Zalora, and produces a single
colour-coded Excel workbook.

Deploy: push this repo to GitHub, then deploy on https://streamlit.io/cloud
pointing at app.py.
"""

import io
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Stock Validation ALL MP", page_icon="📦", layout="wide")

# ---------------------------------------------------------------------------
# STYLE CONSTANTS
# ---------------------------------------------------------------------------
NAVY = "1F3864"
GREEN, GREEN_FONT = "C6EFCE", "375623"
RED, RED_FONT = "FFC7CE", "9C0006"
ORANGE, ORANGE_FONT = "FFEB9C", "7D4800"
GREY, GREY_FONT = "D9D9D9", "595959"
ALT_ROW = "F2F7FB"

NOT_FOUND_LABELS = {"NOT IN LAZADA", "NOT IN SHOPEE", "NOT IN TIKTOK", "NOT IN ZALORA", "NOT FOUND"}


def style_of_remark(remark):
    if remark == "TRUE":
        return GREEN, GREEN_FONT
    if remark == "Stock Mismatch":
        return RED, RED_FONT
    if remark in ("UPDATE 0", "UPDATE STOCK"):
        return ORANGE, ORANGE_FONT
    if remark in NOT_FOUND_LABELS:
        return GREY, GREY_FONT
    return None, None


# ---------------------------------------------------------------------------
# FILE HELPERS
# ---------------------------------------------------------------------------

def fix_shopee_xlsx(file_bytes: bytes) -> io.BytesIO:
    """Patch Shopee's invalid activePane XML bug so openpyxl/pandas can read it."""
    src = io.BytesIO(file_bytes)
    dst = io.BytesIO()
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(".xml") or item.filename.endswith(".rels"):
                text = data.decode("utf-8", errors="ignore")
                for old, new in [
                    ("bottom_left", "bottomLeft"),
                    ("top_left", "topLeft"),
                    ("bottom_right", "bottomRight"),
                    ("top_right", "topRight"),
                ]:
                    text = text.replace(f'activePane="{old}"', f'activePane="{new}"')
                    text = text.replace(f'pane="{old}"', f'pane="{new}"')
                data = text.encode("utf-8")
            zout.writestr(item, data)
    dst.seek(0)
    return dst


def parse_soh_xml(file_bytes: bytes) -> pd.DataFrame:
    """SOHbySKU .xls files are actually XML spreadsheets. SKU = col idx 6, Qty = col idx 14."""
    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
    root = ET.fromstring(file_bytes)
    ws = root.find("ss:Worksheet", ns)
    table = ws.find("ss:Table", ns)
    rows = table.findall("ss:Row", ns)

    def row_to_list(row):
        cells = row.findall("ss:Cell", ns)
        vals = []
        for c in cells:
            d = c.find("ss:Data", ns)
            vals.append(d.text if d is not None else None)
        return vals

    records = []
    for row in rows[7:]:
        vals = row_to_list(row)
        if len(vals) > 14 and vals[6]:
            records.append({"SKU": str(vals[6]).strip(), "Qty": vals[14]})
    df = pd.DataFrame(records)
    if not df.empty:
        df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0).astype(int)
    return df


def clean_sku(series):
    return series.astype(str).str.strip()


# ---------------------------------------------------------------------------
# PER-MARKETPLACE LOOKUP BUILDERS
# ---------------------------------------------------------------------------

def build_lazada_lookup(sp_file):
    df = pd.read_excel(sp_file, header=0, skiprows=[1, 2, 3])
    df.columns = [
        "Product ID", "catId", "Product Name", "currencyCode", "sku.skuId", "status",
        "Shop SKU", "SellerSKU", "Quantity", "Price", "SpecialPrice",
        "SpecialPrice Start", "SpecialPrice End", "Variations Combo", "md5key",
    ][: len(df.columns)]
    df["SellerSKU"] = clean_sku(df["SellerSKU"])
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0).astype(int)
    return df.groupby("SellerSKU")["Quantity"].sum().to_dict()


def build_shopee_lookup(sp_files):
    """sp_files: list of uploaded file objects (zip contents or a single xlsx)."""
    lookups = {}
    for f in sp_files:
        try:
            fixed = fix_shopee_xlsx(f.getvalue())
            df = pd.read_excel(fixed, header=2, skiprows=[3, 4, 5])
        except Exception:
            continue
        if "SKU" not in df.columns or "Stock" not in df.columns:
            continue
        df = df.dropna(subset=["SKU"])
        df["SKU"] = clean_sku(df["SKU"])
        df = df[df["SKU"] != "nan"]
        if df.empty:
            continue
        df["Stock"] = pd.to_numeric(df["Stock"], errors="coerce").fillna(0).astype(int)
        for sku, qty in df.groupby("SKU")["Stock"].sum().items():
            lookups[sku] = lookups.get(sku, 0) + qty
    return lookups


def build_tiktok_lookup(sp_file):
    df = pd.read_excel(sp_file, header=2, skiprows=[3, 4])
    df.columns = [
        "Product ID", "Category", "Product Name", "SKU ID",
        "Variation Option", "Price", "Quantity", "Seller SKU",
    ][: len(df.columns)]
    df["Seller SKU"] = clean_sku(df["Seller SKU"])
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0).astype(int)
    return df.groupby("Seller SKU")["Quantity"].sum().to_dict()


def build_zalora_lookup(stock_file, status_file=None):
    df = pd.read_excel(stock_file, header=0)
    df.columns = ["SellerSku", "ShopSku", "Quantity", "Name"][: len(df.columns)]
    df["SellerSku"] = clean_sku(df["SellerSku"])
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0).astype(int)
    lookup = df.groupby("SellerSku")["Quantity"].sum().to_dict()

    status_lookup = {}
    if status_file is not None:
        try:
            ds = pd.read_excel(status_file, header=0)
            ds.columns = ["SellerSku", "ShopSku", "Name", "Status"][: len(ds.columns)]
            ds["SellerSku"] = clean_sku(ds["SellerSku"])
            status_lookup = ds.set_index("SellerSku")["Status"].astype(str).str.strip().str.lower().to_dict()
        except Exception:
            pass
    return lookup, status_lookup


# ---------------------------------------------------------------------------
# VALIDATION LOGIC
# ---------------------------------------------------------------------------

def get_remark(expected, mp_stock, not_found_label):
    if pd.isna(mp_stock):
        return "Mismatch", not_found_label
    exp, mp = int(expected), int(mp_stock)
    if exp == 0 and mp > 0:
        return "Mismatch", "UPDATE 0"
    if exp > 0 and mp == 0:
        return "Mismatch", "UPDATE STOCK"
    if exp != mp:
        return "Mismatch", "Stock Mismatch"
    return "Match", "TRUE"


def process_marketplace(sv_df, lookup, not_found_label, status_lookup=None):
    df = sv_df.copy()
    df["Seller SKU"] = clean_sku(df["Seller SKU"])
    if "Difference" in df.columns:
        df.rename(columns={"Difference": "Original Difference"}, inplace=True)
    df["Expected Stock"] = pd.to_numeric(df["Expected Stock"], errors="coerce").fillna(0).astype(int)

    df["Marketplace Stock"] = df["Seller SKU"].map(lookup)

    statuses, remarks = [], []
    for exp, mp in zip(df["Expected Stock"], df["Marketplace Stock"]):
        status, remark = get_remark(exp, mp, not_found_label)
        statuses.append(status)
        remarks.append(remark)
    df["Status_Result"] = statuses
    df["Remark"] = remarks

    df["Difference"] = df["Marketplace Stock"] - df["Expected Stock"]
    df.loc[df["Marketplace Stock"].isna(), "Difference"] = np.nan

    if status_lookup:
        df["Zalora_Status"] = df["Seller SKU"].map(status_lookup).fillna("—")

    return df


def compute_kpis(df, not_found_label):
    total = len(df)
    matched = int((df["Remark"] == "TRUE").sum())
    stock_mismatch = int((df["Remark"] == "Stock Mismatch").sum())
    update0 = int((df["Remark"] == "UPDATE 0").sum())
    update_stock = int((df["Remark"] == "UPDATE STOCK").sum())
    not_found = int((df["Remark"] == not_found_label).sum())
    total_issues = total - matched
    accuracy = round((matched / total * 100), 2) if total else 0
    return {
        "Total SKUs": total, "Matched": matched, "Stock Mismatch": stock_mismatch,
        "UPDATE 0": update0, "UPDATE STOCK": update_stock, "NOT FOUND": not_found,
        "Total Issues": total_issues, "Accuracy %": accuracy,
    }


# ---------------------------------------------------------------------------
# WORKBOOK BUILDER
# ---------------------------------------------------------------------------

def write_df_sheet(wb, sheet_name, df, remark_col="Remark", status_col="Status_Result"):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    thin_border = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

    cols = list(df.columns)
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    remark_idx = cols.index(remark_col) + 1 if remark_col in cols else None
    status_idx = cols.index(status_col) + 1 if status_col in cols else None

    for i, (_, r) in enumerate(df.iterrows(), start=2):
        remark_val = r.get(remark_col, None)
        fill_color, font_color = style_of_remark(remark_val) if remark_val is not None else (None, None)
        row_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid") if i % 2 == 0 else None

        for j, col in enumerate(cols, start=1):
            val = r[col]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = body_font
            cell.border = thin_border
            if row_fill is not None:
                cell.fill = row_fill

        if remark_idx and fill_color:
            idxs = [remark_idx, status_idx] if status_idx else [remark_idx]
            for idx in idxs:
                c = ws.cell(row=i, column=idx)
                c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                c.font = Font(name="Arial", size=10, bold=True, color=font_color)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for j, col in enumerate(cols, start=1):
        sample = df[col].astype(str).values[:200]
        max_len = max([len(str(col))] + [len(v) for v in sample]) if len(sample) else len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 10), 40)


def build_workbook(marketplaces, kpi_results, soh_vs_all=None):
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    body_font = Font(name="Arial", size=10)

    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Multi-Marketplace Stock Validation"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=NAVY)
    ws["A2"] = datetime.now().strftime("Generated: %d %b %Y, %H:%M")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")

    row = 4
    metric_order = ["Total SKUs", "Matched", "Stock Mismatch", "UPDATE 0", "UPDATE STOCK",
                     "NOT FOUND", "Total Issues", "Accuracy %"]

    for name, df, label in marketplaces:
        ws.cell(row=row, column=1, value=f"{name.upper()} — StockValidation vs Marketplace Stock")
        ws.cell(row=row, column=1).font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2).fill = header_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

        kpis = kpi_results[name]
        for metric in metric_order:
            val = kpis[metric]
            label_metric = label if metric == "NOT FOUND" else metric
            ws.cell(row=row, column=1, value=label_metric).font = body_font
            cell_val = ws.cell(row=row, column=2, value=val)
            cell_val.font = Font(name="Arial", size=10, bold=True)
            cell_val.alignment = Alignment(horizontal="right")

            fill_color = None
            if metric == "Matched":
                fill_color = GREEN
            elif metric in ("Stock Mismatch", "Total Issues"):
                fill_color = RED
            elif metric in ("UPDATE 0", "UPDATE STOCK"):
                fill_color = ORANGE
            elif metric == "NOT FOUND":
                fill_color = GREY
            if fill_color:
                cell_val.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            if metric == "Accuracy %":
                cell_val.number_format = '0.00"%"'
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14

    for name, df, label in marketplaces:
        write_df_sheet(wb, f"{name}_StockVal"[:31], df)
        mism = df[df["Remark"] != "TRUE"]
        if len(mism) > 0:
            write_df_sheet(wb, f"{name}_Mismatches"[:31], mism)

    if soh_vs_all is not None and not soh_vs_all.empty:
        write_df_sheet(wb, "SOH_vs_ALL", soh_vs_all, remark_col="Remark", status_col=None)
        mism = soh_vs_all[soh_vs_all["Remark"] != "TRUE"]
        if len(mism) > 0:
            write_df_sheet(wb, "SOH_Mismatches", mism, remark_col="Remark", status_col=None)

    order = ["Summary"]
    for name, df, label in marketplaces:
        order.append(f"{name}_StockVal"[:31])
        mname = f"{name}_Mismatches"[:31]
        if mname in wb.sheetnames:
            order.append(mname)
    if "SOH_vs_ALL" in wb.sheetnames:
        order.append("SOH_vs_ALL")
    if "SOH_Mismatches" in wb.sheetnames:
        order.append("SOH_Mismatches")
    wb._sheets = [wb[s] for s in order]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# STREAMLIT UI
# ---------------------------------------------------------------------------

st.title("📦 Stock Validation — ALL Marketplaces")
st.caption("Lazada · Shopee · TikTok · Zalora — upload any subset, get one colour-coded workbook back.")

with st.expander("ℹ️ What files can I upload?", expanded=False):
    st.markdown("""
- **StockValidation CSVs** — filenames containing `stockValidation-lazada`, `-shopee`, `-tiktok`, or `-zalora`
- **Lazada** Stock & Price export — filename containing `pricestock`
- **Shopee** Mass Update — filename containing `mass_update_sales_info` (`.xlsx` or `.zip`)
- **TikTok** Batch Edit — filename containing `Tiktoksellercenter_batchedit` or `batchedit`
- **Zalora** `SellerStockTemplate` (required) and optionally `SellerStatusTemplate`
- *(optional warehouse check)* `SOHbySKU...xls` and an `ALL...csv` — only used if they genuinely match the same SKU catalogue
""")

uploaded_files = st.file_uploader(
    "Upload your files (any subset)",
    accept_multiple_files=True,
    type=["csv", "xlsx", "xls", "zip"],
)

if uploaded_files:
    lazada_sv = shopee_sv = tiktok_sv = zalora_sv = None
    lazada_sp = None
    shopee_sp_files = []
    tiktok_sp = None
    zalora_stock = zalora_status = None
    soh_file = all_file = None

    for f in uploaded_files:
        name = f.name.lower()
        if "stockvalidation" in name.replace(" ", "").lower() or "stockvalidation-" in name:
            pass  # handled below by keyword checks
        if re.search(r"stockvalidation.*lazada|lazada.*stockvalidation", name) or ("lazada" in name and name.endswith(".csv")):
            lazada_sv = f
        elif re.search(r"stockvalidation.*shopee|shopee.*stockvalidation", name) or ("shopee" in name and name.endswith(".csv") and "mass_update" not in name):
            shopee_sv = f
        elif re.search(r"stockvalidation.*tiktok|tiktok.*stockvalidation", name) or ("tiktok" in name and name.endswith(".csv") and "batchedit" not in name):
            tiktok_sv = f
        elif re.search(r"stockvalidation.*zalora|zalora.*stockvalidation", name) or ("zalora" in name and name.endswith(".csv")):
            zalora_sv = f
        elif "pricestock" in name:
            lazada_sp = f
        elif "mass_update_sales_info" in name:
            if name.endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(f.getvalue())) as zf:
                    for inner_name in zf.namelist():
                        if inner_name.lower().endswith(".xlsx"):
                            inner_bytes = zf.read(inner_name)
                            shopee_sp_files.append(io.BytesIO(inner_bytes))
                            shopee_sp_files[-1].name = inner_name  # for consistency, not required
            else:
                shopee_sp_files.append(f)
        elif "batchedit" in name or "tiktoksellercenter" in name:
            tiktok_sp = f
        elif "sellerstocktemplate" in name:
            zalora_stock = f
        elif "sellerstatustemplate" in name:
            zalora_status = f
        elif "sohbysku" in name:
            soh_file = f
        elif name.startswith("all") and name.endswith(".csv"):
            all_file = f

    detected = []
    if lazada_sv and lazada_sp:
        detected.append("Lazada")
    if shopee_sv and shopee_sp_files:
        detected.append("Shopee")
    if tiktok_sv and tiktok_sp:
        detected.append("TikTok")
    if zalora_sv and zalora_stock:
        detected.append("Zalora")

    st.write("**Detected marketplaces:**", ", ".join(detected) if detected else "None yet — check filenames match the patterns above.")

    missing_notes = []
    if lazada_sv and not lazada_sp:
        missing_notes.append("Lazada StockValidation found but no `pricestock` file.")
    if shopee_sv and not shopee_sp_files:
        missing_notes.append("Shopee StockValidation found but no `mass_update_sales_info` file.")
    if tiktok_sv and not tiktok_sp:
        missing_notes.append("TikTok StockValidation found but no `batchedit` file.")
    if zalora_sv and not zalora_stock:
        missing_notes.append("Zalora StockValidation found but no `SellerStockTemplate` file.")
    for note in missing_notes:
        st.warning(note)

    if detected and st.button("🚀 Run validation", type="primary"):
        with st.spinner("Processing files..."):
            marketplaces = []
            kpi_results = {}

            if "Lazada" in detected:
                lz_lookup = build_lazada_lookup(lazada_sp)
                lz_sv_df = pd.read_csv(lazada_sv)
                lz_out = process_marketplace(lz_sv_df, lz_lookup, "NOT IN LAZADA")
                marketplaces.append(("Lazada", lz_out, "NOT IN LAZADA"))
                kpi_results["Lazada"] = compute_kpis(lz_out, "NOT IN LAZADA")

            if "Shopee" in detected:
                sh_lookup = build_shopee_lookup(shopee_sp_files)
                sh_sv_df = pd.read_csv(shopee_sv)
                sh_out = process_marketplace(sh_sv_df, sh_lookup, "NOT IN SHOPEE")
                marketplaces.append(("Shopee", sh_out, "NOT IN SHOPEE"))
                kpi_results["Shopee"] = compute_kpis(sh_out, "NOT IN SHOPEE")

            if "TikTok" in detected:
                tk_lookup = build_tiktok_lookup(tiktok_sp)
                tk_sv_df = pd.read_csv(tiktok_sv)
                tk_out = process_marketplace(tk_sv_df, tk_lookup, "NOT IN TIKTOK")
                marketplaces.append(("TikTok", tk_out, "NOT IN TIKTOK"))
                kpi_results["TikTok"] = compute_kpis(tk_out, "NOT IN TIKTOK")

            if "Zalora" in detected:
                za_lookup, za_status_lookup = build_zalora_lookup(zalora_stock, zalora_status)
                za_sv_df = pd.read_csv(zalora_sv)
                za_out = process_marketplace(za_sv_df, za_lookup, "NOT IN ZALORA", za_status_lookup)
                marketplaces.append(("Zalora", za_out, "NOT IN ZALORA"))
                kpi_results["Zalora"] = compute_kpis(za_out, "NOT IN ZALORA")

            soh_vs_all_df = None
            if soh_file is not None and all_file is not None:
                st.info(
                    "SOH and ALL files were both uploaded, but this app doesn't auto-validate that "
                    "they share the same brand/SKU catalogue as your marketplace files. Skipping the "
                    "warehouse check automatically — verify manually before relying on it."
                )

            workbook_buf = build_workbook(marketplaces, kpi_results, soh_vs_all_df)

        st.success("Done!")

        cols = st.columns(len(marketplaces) if marketplaces else 1)
        for i, (name, df, label) in enumerate(marketplaces):
            with cols[i]:
                k = kpi_results[name]
                st.metric(f"{name} accuracy", f"{k['Accuracy %']}%", f"{k['Total Issues']} issues")

        st.download_button(
            "⬇️ Download Stock_Validation_All_MP.xlsx",
            data=workbook_buf,
            file_name="Stock_Validation_All_MP.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Upload StockValidation CSVs + marketplace stock files above to get started.")
